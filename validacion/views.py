from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from datetime import datetime, timedelta
from django.db.models import Q
from django.contrib.auth.models import User
from accounts.models import User  # ← ESTA ES LA BUENA
from django.core.paginator import Paginator
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import csv
import io
from .models import Pago, Incidencia, Bitacora, Conciliacion, TipoAccion, AlertaAdmin, TablaSistema
from .forms import PagoForm
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.contrib.auth.decorators import login_required
from django.contrib import messages

@login_required(login_url='/accounts/login/')
def dashboard(request):
    hoy = timezone.now().date()
    
    # KPIs DINÁMICOS
    total_pagos_hoy = Pago.objects.filter(created_at__date=hoy).count()
    incidencias_abiertas = Incidencia.objects.filter(estado='ABIERTA').count()
    firmas_pendientes = Incidencia.objects.filter(estado='PENDIENTE_FIRMA').count()
    
    # Alertas recientes (últimas 24 horas)
    alertas_recientes = Incidencia.objects.filter(
        fecha_apertura__gte=hoy - timedelta(days=1)
    ).order_by('-fecha_apertura')[:3]
    
    # Pendientes de firma
    pendientes_firma = Incidencia.objects.filter(
        estado='PENDIENTE_FIRMA'
    ).order_by('-fecha_apertura')[:3]
    
    # Actividad reciente (bitácora) - Formatear fecha completa
    actividad_bitacora = Bitacora.objects.all().order_by('-fecha')[:3]
    actividad_reciente = []
    for actividad in actividad_bitacora:
        actividad_reciente.append({
            'fecha_hora': actividad.fecha.strftime('%d/%m/%Y %H:%M:%S'),
            'usuario': actividad.usuario.username,
            'accion': actividad.tipo_accion.nombre,
            'objeto': actividad.tabla_sistema.nombre
        })
    
    context = {
        'total_pagos_hoy': total_pagos_hoy,
        'incidencias_abiertas': incidencias_abiertas,
        'firmas_pendientes': firmas_pendientes,
        'alertas_recientes': alertas_recientes,
        'pendientes_firma': pendientes_firma,
        'actividad_reciente': actividad_reciente,
    }
    
    return render(request, 'dashboard.html', context)

# VISTA BITÁCORA COMPLETA
@login_required(login_url='/accounts/login/')
def bitacora(request):
    # Obtener parámetros de filtro
    fecha_desde = request.GET.get('desde', '')
    fecha_hasta = request.GET.get('hasta', '')
    usuario_filtro = request.GET.get('usuario', '')
    accion_filtro = request.GET.get('accion', '')
    texto_libre = request.GET.get('q', '')
    pagina = request.GET.get('pagina', 1)
    
    # Consulta base
    bitacoras = Bitacora.objects.all().select_related('usuario', 'tipo_accion', 'tabla_sistema').order_by('-fecha')
    
    # Aplicar filtros
    if fecha_desde:
        bitacoras = bitacoras.filter(fecha__date__gte=fecha_desde)
    if fecha_hasta:
        bitacoras = bitacoras.filter(fecha__date__lte=fecha_hasta)
    if usuario_filtro and usuario_filtro != 'Todos':
        bitacoras = bitacoras.filter(usuario__username=usuario_filtro)
    if accion_filtro and accion_filtro != 'Todas':
        bitacoras = bitacoras.filter(tipo_accion__nombre=accion_filtro)
    if texto_libre:
        bitacoras = bitacoras.filter(
            Q(observacion__icontains=texto_libre) |
            Q(valores_nuevos__icontains=texto_libre) |
            Q(usuario__username__icontains=texto_libre) |
            Q(tabla_sistema__nombre__icontains=texto_libre)
        )
    
    # ✅ NUEVO: Formatear registros con fecha completa
    registros_formateados = []
    for bitacora in bitacoras:
        registros_formateados.append({
            'fecha_hora': bitacora.fecha.strftime('%d/%m/%Y %H:%M:%S'),  # Fecha completa
            'usuario': bitacora.usuario.username,
            'accion': bitacora.tipo_accion.nombre,
            'objeto': bitacora.tabla_sistema.nombre,
            'observacion': bitacora.observacion,
            'registro_afectado': bitacora.registro_afectado,
            'valores_nuevos': bitacora.valores_nuevos
        })
    
    # Paginación con registros formateados
    paginator = Paginator(registros_formateados, 20)
    page_obj = paginator.get_page(pagina)
    
    # Estadísticas (usar la consulta original para counts)
    total_registros = bitacoras.count()
    validaciones = bitacoras.filter(tipo_accion__nombre__icontains='VALIDAR').count()
    anulaciones = bitacoras.filter(tipo_accion__nombre__icontains='ANULAR').count()
    alertas = bitacoras.filter(tipo_accion__nombre__icontains='ALERTA').count()
    
    # Obtener opciones para los filtros
    usuarios = User.objects.filter(bitacora__isnull=False).distinct()
    acciones = TipoAccion.objects.all()
    
    context = {
        'bitacoras': page_obj,  # Ahora con fecha formateada
        'total_registros': total_registros,
        'validaciones': validaciones,
        'anulaciones': anulaciones,
        'alertas': alertas,
        'usuarios': usuarios,
        'acciones': acciones,
        'filtros': {
            'desde': fecha_desde,
            'hasta': fecha_hasta,
            'usuario': usuario_filtro,
            'accion': accion_filtro,
            'texto': texto_libre,
        }
    }
    
    return render(request, 'bitacora.html', context)

# EXPORTAR BITÁCORA A PDF
@login_required(login_url='/accounts/login/')
def exportar_bitacora_pdf(request):
    # Obtener los mismos filtros que en la vista bitacora
    fecha_desde = request.GET.get('desde', '')
    fecha_hasta = request.GET.get('hasta', '')
    usuario_filtro = request.GET.get('usuario', '')
    accion_filtro = request.GET.get('accion', '')
    texto_libre = request.GET.get('q', '')
    
    # Aplicar los mismos filtros
    bitacoras = Bitacora.objects.all().select_related('usuario', 'tipo_accion', 'tabla_sistema').order_by('-fecha')
    
    if fecha_desde:
        bitacoras = bitacoras.filter(fecha__date__gte=fecha_desde)
    if fecha_hasta:
        bitacoras = bitacoras.filter(fecha__date__lte=fecha_hasta)
    if usuario_filtro and usuario_filtro != 'Todos':
        bitacoras = bitacoras.filter(usuario__username=usuario_filtro)
    if accion_filtro and accion_filtro != 'Todas':
        bitacoras = bitacoras.filter(tipo_accion__nombre=accion_filtro)
    if texto_libre:
        bitacoras = bitacoras.filter(
            Q(observacion__icontains=texto_libre) |
            Q(valores_nuevos__icontains=texto_libre) |
            Q(usuario__username__icontains=texto_libre) |
            Q(tabla_sistema__nombre__icontains=texto_libre)
        )
    
    # Crear el PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    
    # Título
    title = Paragraph("Bitácora del Sistema - ISDM Validación", styles['Title'])
    elements.append(title)
    
    # Información de filtros
    filtros_text = f"Filtros aplicados: Desde {fecha_desde if fecha_desde else 'N/A'} - Hasta {fecha_hasta if fecha_hasta else 'N/A'} - Usuario: {usuario_filtro if usuario_filtro else 'Todos'} - Acción: {accion_filtro if accion_filtro else 'Todas'}"
    filtros_para = Paragraph(filtros_text, styles['Normal'])
    elements.append(filtros_para)
    
    elements.append(Paragraph("<br/>", styles['Normal']))
    
    # Preparar datos para la tabla
    data = [['Fecha/Hora', 'Usuario', 'Acción', 'Objeto', 'Observación']]
    
    for bitacora in bitacoras:
        fecha_str = bitacora.fecha.strftime('%d/%m/%Y %H:%M:%S')
        data.append([
            fecha_str,
            bitacora.usuario.username,
            bitacora.tipo_accion.nombre,
            bitacora.tabla_sistema.nombre,
            bitacora.observacion[:50] + '...' if len(bitacora.observacion) > 50 else bitacora.observacion
        ])
    
    # Crear tabla
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343a40')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
    ]))
    
    elements.append(table)
    
    # Construir PDF
    doc.build(elements)
    
    # Preparar respuesta
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="bitacora_isdm.pdf"'
    
    return response

# EXPORTAR BITÁCORA A EXCEL (CSV)
@login_required(login_url='/accounts/login/')
def exportar_bitacora_excel(request):
    # Obtener los mismos filtros que en la vista bitacora
    fecha_desde = request.GET.get('desde', '')
    fecha_hasta = request.GET.get('hasta', '')
    usuario_filtro = request.GET.get('usuario', '')
    accion_filtro = request.GET.get('accion', '')
    texto_libre = request.GET.get('q', '')
    
    # Aplicar los mismos filtros
    bitacoras = Bitacora.objects.all().select_related('usuario', 'tipo_accion', 'tabla_sistema').order_by('-fecha')
    
    if fecha_desde:
        bitacoras = bitacoras.filter(fecha__date__gte=fecha_desde)
    if fecha_hasta:
        bitacoras = bitacoras.filter(fecha__date__lte=fecha_hasta)
    if usuario_filtro and usuario_filtro != 'Todos':
        bitacoras = bitacoras.filter(usuario__username=usuario_filtro)
    if accion_filtro and accion_filtro != 'Todas':
        bitacoras = bitacoras.filter(tipo_accion__nombre=accion_filtro)
    if texto_libre:
        bitacoras = bitacoras.filter(
            Q(observacion__icontains=texto_libre) |
            Q(valores_nuevos__icontains=texto_libre) |
            Q(usuario__username__icontains=texto_libre) |
            Q(tabla_sistema__nombre__icontains=texto_libre)
        )
    
    # Crear respuesta CSV (simulando Excel)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="bitacora_isdm.csv"'
    
    writer = csv.writer(response)
    # Escribir encabezados
    writer.writerow(['Fecha/Hora', 'Usuario', 'Acción', 'Objeto', 'Observación', 'Valores Nuevos', 'Registro Afectado'])
    
    # Escribir datos
    for bitacora in bitacoras:
        fecha_str = bitacora.fecha.strftime('%d/%m/%Y %H:%M:%S')
        writer.writerow([
            fecha_str,
            bitacora.usuario.username,
            bitacora.tipo_accion.nombre,
            bitacora.tabla_sistema.nombre,
            bitacora.observacion,
            str(bitacora.valores_nuevos) if bitacora.valores_nuevos else '',
            bitacora.registro_afectado
        ])
    
    return response

# VISTA CONCILIACIÓN COMPLETA
@login_required(login_url='/accounts/login/')
def conciliacion(request):
    # Obtener parámetros del formulario
    fecha_conciliacion = request.GET.get('fecha', '')
    pasarela_filtro = request.GET.get('pasarela', 'MacroClick')
    
    # Consulta base para pagos del sistema
    pagos_sistema = Pago.objects.all().order_by('-created_at')
    
    # Si hay fecha, filtrar por fecha
    if fecha_conciliacion:
        pagos_sistema = pagos_sistema.filter(created_at__date=fecha_conciliacion)
    
    # Simular datos externos (esto luego vendrá de archivos CSV)
    datos_externos = [
        {'referencia': '#P-00120', 'monto': 10000, 'origen': 'Tarjeta', 'estado': 'OK'},
        {'referencia': '#P-00121', 'monto': 10000, 'origen': 'Tarjeta', 'estado': 'DIF'},
        {'referencia': '', 'monto': '', 'origen': '', 'estado': 'SIN_MATCH'},
    ]
    
    # Procesar matching y estados
    for pago in pagos_sistema:
        pago.estado_conciliacion = 'SIN_MATCH'  # Por defecto
        pago.monto_externo = None
        
        for externo in datos_externos:
            if externo['referencia'] == pago.referencia:
                if externo['monto'] == float(pago.monto):
                    pago.estado_conciliacion = 'OK'
                else:
                    pago.estado_conciliacion = 'DIF'
                    pago.monto_externo = externo['monto']
                break
    
    # Estadísticas
    total_pagos = pagos_sistema.count()
    pagos_ok = sum(1 for p in pagos_sistema if getattr(p, 'estado_conciliacion', '') == 'OK')
    pagos_dif = sum(1 for p in pagos_sistema if getattr(p, 'estado_conciliacion', '') == 'DIF')
    pagos_sin_match = sum(1 for p in pagos_sistema if getattr(p, 'estado_conciliacion', '') == 'SIN_MATCH')
    
    context = {
        'pagos_sistema': pagos_sistema,
        'datos_externos': datos_externos,
        'total_pagos': total_pagos,
        'pagos_ok': pagos_ok,
        'pagos_dif': pagos_dif,
        'pagos_sin_match': pagos_sin_match,
        'filtros': {
            'fecha': fecha_conciliacion,
            'pasarela': pasarela_filtro,
        }
    }
    
    return render(request, 'conciliacion.html', context)

@login_required(login_url='/accounts/login/')
def subir_reporte_externo(request):
    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']
        # Aquí procesarías el archivo CSV y guardarías los datos
        # Por ahora solo redirigimos
        return redirect('conciliacion')
    return redirect('conciliacion')

@login_required(login_url='/accounts/login/')
def generar_incidencias(request):
    if request.method == 'POST':
        # Lógica para generar incidencias automáticamente
        pagos_sistema = Pago.objects.all()
        
        for pago in pagos_sistema:
            # Simular lógica de detección de diferencias
            if pago.referencia == '#P-00121':  # Ejemplo de pago con diferencia
                Incidencia.objects.create(
                    numero_referencia=f"DIF-{pago.referencia}",
                    tipo_incidencia='DIF_MONTO',
                    descripcion=f"Diferencia detectada en {pago.referencia}",
                    pago_relacionado=pago,
                    monto_diferencia=200  # Ejemplo: 10,000 vs 9,800
                )
            elif pago.referencia == '#P-00122':  # Ejemplo de pago sin match
                Incidencia.objects.create(
                    numero_referencia=f"SM-{pago.referencia}",
                    tipo_incidencia='SIN_MATCH',
                    descripcion=f"Pago sin match externo: {pago.referencia}",
                    pago_relacionado=pago
                )
        
        return redirect('conciliacion')
    return redirect('conciliacion')

@login_required(login_url='/accounts/login/')
def lista_pagos(request):
    # Obtener parámetros de filtro
    fecha_desde = request.GET.get('desde', '')
    fecha_hasta = request.GET.get('hasta', '')
    estado_filtro = request.GET.get('estado', 'Todos')
    metodo_filtro = request.GET.get('metodo', 'Todos')
    texto_libre = request.GET.get('q', '')
    pagina = request.GET.get('pagina', 1)
    
    # Consulta base
    pagos = Pago.objects.all().order_by('-created_at')
    
    # Aplicar filtros
    if fecha_desde:
        pagos = pagos.filter(fecha_pago__date__gte=fecha_desde)
    if fecha_hasta:
        pagos = pagos.filter(fecha_pago__date__lte=fecha_hasta)
    if estado_filtro and estado_filtro != 'Todos':
        pagos = pagos.filter(estado=estado_filtro)
    if metodo_filtro and metodo_filtro != 'Todos':
        pagos = pagos.filter(metodo_pago=metodo_filtro)
    if texto_libre:
        pagos = pagos.filter(
            Q(referencia__icontains=texto_libre) |
            Q(estudiante_nombre__icontains=texto_libre) |
            Q(concepto__icontains=texto_libre)
        )
    
    # Paginación
    paginator = Paginator(pagos, 20)
    page_obj = paginator.get_page(pagina)
    
    # Estadísticas para las tarjetas (usar consulta filtrada)
    total_pagos = pagos.count()
    pagos_pagados = pagos.filter(estado='PAGADO').count()
    pagos_pendientes = pagos.filter(estado='PENDIENTE').count()
    pagos_vencidos = pagos.filter(estado='VENCIDO').count()
    
    context = {
        'pagos': page_obj,
        'total_pagos': total_pagos,
        'pagos_pagados': pagos_pagados,
        'pagos_pendientes': pagos_pendientes,
        'pagos_vencidos': pagos_vencidos,
        'filtros': {
            'desde': fecha_desde,
            'hasta': fecha_hasta,
            'estado': estado_filtro,
            'metodo': metodo_filtro,
            'texto': texto_libre,
        }
    }
    
    return render(request, 'lista_pagos.html', context)

# NUEVA FUNCIÓN PARA AGREGAR PAGOS
@login_required(login_url='/accounts/login/')
def agregar_pago(request):
    if request.method == 'POST':
        form = PagoForm(request.POST, request.FILES)
        if form.is_valid():
            pago = form.save(commit=False)
            pago.usuario_creacion = request.user
            
            # Combinar fecha y hora en los campos DateTime
            fecha_pago = form.cleaned_data['fecha_pago']
            hora_pago = form.cleaned_data['hora_pago']
            pago.fecha_pago = datetime.combine(fecha_pago, hora_pago)
            
            fecha_vencimiento = form.cleaned_data['fecha_vencimiento']
            hora_vencimiento = form.cleaned_data['hora_vencimiento']
            pago.fecha_vencimiento = datetime.combine(fecha_vencimiento, hora_vencimiento)
            
            pago.save()
            
            # Registrar en bitácora (si las tablas existen)
            try:
                tipo_accion, created = TipoAccion.objects.get_or_create(
                    nombre='CREAR_PAGO',
                    defaults={'estado_id': 1, 'tabla_sistema_id': 1}
                )
                tabla_sistema, created = TablaSistema.objects.get_or_create(
                    nombre='Pagos',
                    defaults={'descripcion': 'Tabla de pagos del sistema', 'importancia': 1, 'estado_id': 1}
                )
                
                Bitacora.objects.create(
                    tipo_accion=tipo_accion,
                    tabla_sistema=tabla_sistema,
                    observacion=f"Pago creado: {pago.referencia} - {pago.estudiante_nombre} - ${pago.monto}",
                    usuario=request.user,
                    registro_afectado=pago.id
                )
            except Exception as e:
                # Si hay error con la bitácora, continuar igual
                print(f"Error al registrar en bitácora: {e}")
            
            # Determinar a dónde redirigir según el botón presionado
            if 'guardar_agregar_otro' in request.POST:
                return redirect('agregar_pago')
            elif 'guardar_continuar' in request.POST:
                # En un futuro podrías crear una vista de edición
                return redirect('lista_pagos')
            else:
                return redirect('lista_pagos')
    else:
        form = PagoForm()
    
    # ✅ CORREGIDO: Usar 'pago.html' en lugar de 'pagos_cuotas.html'
    return render(request, 'pago.html', {'form': form})

@login_required(login_url='/accounts/login/')
def exportar_pagos_pdf(request):
    # Obtener parámetros de filtro (igual que en lista_pagos)
    fecha_desde = request.GET.get('desde', '')
    fecha_hasta = request.GET.get('hasta', '')
    estado_filtro = request.GET.get('estado', 'Todos')
    metodo_filtro = request.GET.get('metodo', 'Todos')
    texto_libre = request.GET.get('q', '')
    
    # Consulta base
    pagos = Pago.objects.all().order_by('-created_at')
    
    # Aplicar filtros
    if fecha_desde:
        pagos = pagos.filter(fecha_pago__date__gte=fecha_desde)
    if fecha_hasta:
        pagos = pagos.filter(fecha_pago__date__lte=fecha_hasta)
    if estado_filtro and estado_filtro != 'Todos':
        pagos = pagos.filter(estado=estado_filtro)
    if metodo_filtro and metodo_filtro != 'Todos':
        pagos = pagos.filter(metodo_pago=metodo_filtro)
    if texto_libre:
        pagos = pagos.filter(
            Q(referencia__icontains=texto_libre) |
            Q(estudiante_nombre__icontains=texto_libre) |
            Q(concepto__icontains=texto_libre)
        )
    
    # Crear el PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    
    # Título
    title = Paragraph("Reporte de Pagos - ISDM Validación", styles['Title'])
    elements.append(title)
    
    # Información de filtros
    filtros_lista = []
    if fecha_desde: filtros_lista.append(f"Desde: {fecha_desde}")
    if fecha_hasta: filtros_lista.append(f"Hasta: {fecha_hasta}")
    if estado_filtro != 'Todos': filtros_lista.append(f"Estado: {estado_filtro}")
    if metodo_filtro != 'Todos': filtros_lista.append(f"Método: {metodo_filtro}")
    if texto_libre: filtros_lista.append(f"Búsqueda: {texto_libre}")
    
    if filtros_lista:
        filtros_text = "Filtros aplicados: " + " | ".join(filtros_lista)
        filtros_para = Paragraph(filtros_text, styles['Normal'])
        elements.append(filtros_para)
    
    elements.append(Paragraph("<br/>", styles['Normal']))
    
    # Preparar datos para la tabla
    data = [['Referencia', 'Estudiante', 'Monto', 'Fecha Pago', 'Concepto', 'Estado', 'Método']]
    
    for pago in pagos:
        data.append([
            pago.referencia,
            pago.estudiante_nombre,
            f"${pago.monto:,.2f}",
            pago.fecha_pago.strftime('%d/%m/%Y %H:%M') if pago.fecha_pago else '',
            pago.concepto,
            pago.estado,
            pago.metodo_pago
        ])
    
    # Crear tabla
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343a40')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
    ]))
    
    elements.append(table)
    
    # Estadísticas resumen
    elements.append(Paragraph("<br/>", styles['Normal']))
    total_monto = sum(pago.monto for pago in pagos)
    resumen_text = f"Total registros: {pagos.count()} | Monto total: ${total_monto:,.2f}"
    resumen_para = Paragraph(resumen_text, styles['Normal'])
    elements.append(resumen_para)
    
    # Pie de página
    elements.append(Paragraph(f"<br/>Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    
    # Construir PDF
    doc.build(elements)
    
    # Preparar respuesta
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_pagos.pdf"'
    
    return response

@login_required(login_url='/accounts/login/')
def exportar_pagos_excel(request):
    # Obtener parámetros de filtro (igual que en lista_pagos)
    fecha_desde = request.GET.get('desde', '')
    fecha_hasta = request.GET.get('hasta', '')
    estado_filtro = request.GET.get('estado', 'Todos')
    metodo_filtro = request.GET.get('metodo', 'Todos')
    texto_libre = request.GET.get('q', '')
    
    # Consulta base
    pagos = Pago.objects.all().order_by('-created_at')
    
    # Aplicar filtros
    if fecha_desde:
        pagos = pagos.filter(fecha_pago__date__gte=fecha_desde)
    if fecha_hasta:
        pagos = pagos.filter(fecha_pago__date__lte=fecha_hasta)
    if estado_filtro and estado_filtro != 'Todos':
        pagos = pagos.filter(estado=estado_filtro)
    if metodo_filtro and metodo_filtro != 'Todos':
        pagos = pagos.filter(metodo_pago=metodo_filtro)
    if texto_libre:
        pagos = pagos.filter(
            Q(referencia__icontains=texto_libre) |
            Q(estudiante_nombre__icontains=texto_libre) |
            Q(concepto__icontains=texto_libre)
        )
    
    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pagos"
    
    # Título
    ws.merge_cells('A1:G1')
    ws['A1'] = "Reporte de Pagos - ISDM Validación"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Información de filtros
    row_num = 3
    filtros_lista = []
    if fecha_desde: filtros_lista.append(f"Desde: {fecha_desde}")
    if fecha_hasta: filtros_lista.append(f"Hasta: {fecha_hasta}")
    if estado_filtro != 'Todos': filtros_lista.append(f"Estado: {estado_filtro}")
    if metodo_filtro != 'Todos': filtros_lista.append(f"Método: {metodo_filtro}")
    if texto_libre: filtros_lista.append(f"Búsqueda: {texto_libre}")
    
    if filtros_lista:
        ws.merge_cells(f'A{row_num}:G{row_num}')
        ws[f'A{row_num}'] = "Filtros aplicados: " + " | ".join(filtros_lista)
        ws[f'A{row_num}'].font = Font(italic=True)
        row_num += 2
    
    # Encabezados de tabla
    headers = ['Referencia', 'Estudiante', 'Monto', 'Fecha Pago', 'Concepto', 'Estado', 'Método']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='343a40', end_color='343a40', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
    
    # Datos
    for pago in pagos:
        row_num += 1
        ws.cell(row=row_num, column=1, value=pago.referencia)
        ws.cell(row=row_num, column=2, value=pago.estudiante_nombre)
        ws.cell(row=row_num, column=3, value=float(pago.monto))
        ws.cell(row=row_num, column=4, value=pago.fecha_pago.strftime('%d/%m/%Y %H:%M') if pago.fecha_pago else '')
        ws.cell(row=row_num, column=5, value=pago.concepto)
        ws.cell(row=row_num, column=6, value=pago.estado)
        ws.cell(row=row_num, column=7, value=pago.metodo_pago)
    
    # Formato de moneda para la columna de monto
    for row in range(row_num - pagos.count() + 1, row_num + 1):
        ws.cell(row=row, column=3).number_format = '"$"#,##0.00'
    
    # Estadísticas resumen
    row_num += 2
    total_monto = sum(pago.monto for pago in pagos)
    ws.merge_cells(f'A{row_num}:G{row_num}')
    ws[f'A{row_num}'] = f"Total registros: {pagos.count()} | Monto total: ${total_monto:,.2f}"
    ws[f'A{row_num}'].font = Font(bold=True)
    
    # Fecha de generación
    row_num += 1
    ws.merge_cells(f'A{row_num}:G{row_num}')
    ws[f'A{row_num}'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws[f'A{row_num}'].font = Font(italic=True)
    
    # Ajustar anchos de columna
    column_widths = {
        'A': 20,  # Referencia
        'B': 25,  # Estudiante
        'C': 15,  # Monto
        'D': 20,  # Fecha Pago
        'E': 30,  # Concepto
        'F': 15,  # Estado
        'G': 15   # Método
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Crear respuesta
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_pagos.xlsx"'
    
    return response

# VISTA PARA VER DETALLE DE PAGO - CORREGIDA
@login_required(login_url='/accounts/login/')
def detalle_pago(request, pago_id):
    """Vista para ver detalles de un pago específico"""
    pago = get_object_or_404(Pago, id=pago_id)
    
    # Obtener incidencias relacionadas con este pago
    incidencias_relacionadas = Incidencia.objects.filter(pago_relacionado=pago)
    
    # Obtener actividad de bitácora relacionada
    actividad_pago = Bitacora.objects.filter(registro_afectado=pago.id).order_by('-fecha')
    
    # Pasar los tipos de incidencia al template
    tipos_incidencia = Incidencia.TIPOS_INCIDENCIA
    
    context = {
        'pago': pago,
        'incidencias_relacionadas': incidencias_relacionadas,
        'actividad_pago': actividad_pago,
        'tipos_incidencia': tipos_incidencia,  # NUEVO: agregar esto
        'titulo': f'Detalle - {pago.referencia}',
        'now': timezone.now()
    }
    return render(request, 'detalle_pago.html', context)

# VISTA PARA EDITAR PAGO - CORREGIDA
@login_required(login_url='/accounts/login/')
def editar_pago(request, pago_id):
    """Vista para editar un pago existente"""
    pago = get_object_or_404(Pago, id=pago_id)
    
    if request.method == 'POST':
        form = PagoForm(request.POST, request.FILES, instance=pago)
        if form.is_valid():
            pago_editado = form.save(commit=False)
            
            # Combinar fecha y hora (si tu form tiene campos separados)
            if 'fecha_pago' in form.cleaned_data and 'hora_pago' in form.cleaned_data:
                fecha_pago = form.cleaned_data['fecha_pago']
                hora_pago = form.cleaned_data['hora_pago']
                pago_editado.fecha_pago = datetime.combine(fecha_pago, hora_pago)
            
            pago_editado.save()
            
            # Registrar en bitácora
            try:
                tipo_accion, created = TipoAccion.objects.get_or_create(
                    nombre='MODIFICAR_PAGO',
                    defaults={'estado_id': 1, 'tabla_sistema_id': 1}
                )
                tabla_sistema, created = TablaSistema.objects.get_or_create(
                    nombre='Pagos',
                    defaults={'descripcion': 'Tabla de pagos del sistema', 'importancia': 1, 'estado_id': 1}
                )
                
                Bitacora.objects.create(
                    tipo_accion=tipo_accion,
                    tabla_sistema=tabla_sistema,
                    observacion=f"Pago modificado: {pago.referencia}",
                    usuario=request.user,
                    registro_afectado=pago.id,
                    valores_nuevos={
                        'referencia': pago.referencia,
                        'estudiante': pago.estudiante_nombre,
                        'monto': str(pago.monto),
                        'estado': pago.estado
                    }
                )
            except Exception as e:
                print(f"Error al registrar en bitácora: {e}")
            
            messages.success(request, f'✅ Pago {pago.referencia} actualizado correctamente')
            return redirect('lista_pagos')
    else:
        form = PagoForm(instance=pago)
    
    context = {
        'form': form,
        'pago': pago,
        'titulo': f'Editar - {pago.referencia}'
    }
    # ✅ CORRECCIÓN: Usar la ruta correcta del template
    return render(request, 'editar_pago.html', context)

# VISTA PARA ELIMINAR PAGO (OPCIONAL)
@login_required(login_url='/accounts/login/')
def eliminar_pago(request, pago_id):
    """Vista para eliminar un pago"""
    pago = get_object_or_404(Pago, id=pago_id)
    
    if request.method == 'POST':
        referencia = pago.referencia
        pago.delete()
        
        # Registrar en bitácora
        try:
            tipo_accion, created = TipoAccion.objects.get_or_create(
                nombre='ELIMINAR_PAGO',
                defaults={'estado_id': 1, 'tabla_sistema_id': 1}
            )
            tabla_sistema, created = TablaSistema.objects.get_or_create(
                nombre='Pagos',
                defaults={'descripcion': 'Tabla de pagos del sistema', 'importancia': 1, 'estado_id': 1}
            )
            
            Bitacora.objects.create(
                tipo_accion=tipo_accion,
                tabla_sistema=tabla_sistema,
                observacion=f"Pago eliminado: {referencia}",
                usuario=request.user,
                registro_afectado=pago_id
            )
        except Exception as e:
            print(f"Error al registrar en bitácora: {e}")
        
        messages.success(request, f'✅ Pago {referencia} eliminado correctamente')
        return redirect('lista_pagos')
    
    context = {
        'pago': pago
    }
    return render(request, 'eliminar_pago.html', context)

@login_required(login_url='/accounts/login/')
def crear_incidencia(request, pago_id):
    """Vista para crear una incidencia desde el detalle de pago"""
    pago = get_object_or_404(Pago, id=pago_id)
    
    if request.method == 'POST':
        try:
            # Generar número de referencia único para la incidencia
            ultima_incidencia = Incidencia.objects.order_by('-id').first()
            if ultima_incidencia:
                try:
                    ultimo_numero = int(ultima_incidencia.numero_referencia.split('-')[1])
                    nuevo_numero = f"#I-{ultimo_numero + 1:05d}"
                except (ValueError, IndexError):
                    nuevo_numero = "#I-00001"
            else:
                nuevo_numero = "#I-00001"
            
            # Crear la incidencia
            incidencia = Incidencia(
                numero_referencia=nuevo_numero,
                tipo_incidencia=request.POST.get('tipo_incidencia'),
                descripcion=request.POST.get('descripcion'),
                justificacion=request.POST.get('justificacion', ''),
                pago_relacionado=pago,
                usuario_asignado=request.user,
            )
            
            # Manejar monto diferencia si existe
            monto_diferencia = request.POST.get('monto_diferencia')
            if monto_diferencia and monto_diferencia.strip():
                incidencia.monto_diferencia = monto_diferencia
            
            incidencia.save()
            
            # Registrar en bitácora
            try:
                tipo_accion, created = TipoAccion.objects.get_or_create(
                    nombre='CREAR_INCIDENCIA',
                    defaults={'estado_id': 1, 'tabla_sistema_id': 1}
                )
                tabla_sistema, created = TablaSistema.objects.get_or_create(
                    nombre='Incidencias',
                    defaults={'descripcion': 'Tabla de incidencias del sistema', 'importancia': 1, 'estado_id': 1}
                )
                
                Bitacora.objects.create(
                    tipo_accion=tipo_accion,
                    tabla_sistema=tabla_sistema,
                    observacion=f"Incidencia {nuevo_numero} creada para pago {pago.referencia}",
                    usuario=request.user,
                    registro_afectado=incidencia.id,
                    valores_nuevos={
                        'tipo_incidencia': incidencia.tipo_incidencia,
                        'pago_relacionado': pago.referencia,
                        'descripcion': incidencia.descripcion[:100] + '...' if len(incidencia.descripcion) > 100 else incidencia.descripcion
                    }
                )
            except Exception as e:
                print(f"Error en bitácora: {e}")
            
            messages.success(request, f'✅ Incidencia {nuevo_numero} creada exitosamente')
            return redirect('detalle_pago', pago_id=pago.id)
            
        except Exception as e:
            messages.error(request, f'❌ Error al crear la incidencia: {str(e)}')
            return redirect('detalle_pago', pago_id=pago.id)
    
    # Si no es POST, redirigir al detalle
    return redirect('detalle_pago', pago_id=pago.id)

@login_required(login_url='/accounts/login/')
def generar_arqueo(request):
    if request.method == 'POST':
        hoy = timezone.now().date()
        
        try:
            # Calcular métricas para el arqueo
            total_pagos = Pago.objects.filter(created_at__date=hoy).count()
            
            # Usar los mismos cálculos que en el dashboard
            pagos_ok = sum(1 for p in Pago.objects.filter(created_at__date=hoy) 
                          if getattr(p, 'estado_conciliacion', '') == 'OK')
            pagos_dif = sum(1 for p in Pago.objects.filter(created_at__date=hoy) 
                           if getattr(p, 'estado_conciliacion', '') == 'DIF')
            pagos_sin_match = sum(1 for p in Pago.objects.filter(created_at__date=hoy) 
                                 if getattr(p, 'estado_conciliacion', '') == 'SIN_MATCH')
            
            # Crear registro de conciliación
            conciliacion = Conciliacion.objects.create(
                fecha_conciliacion=hoy,
                total_registros=total_pagos,
                match_sistema=(pagos_ok / total_pagos * 100) if total_pagos > 0 else 0,
                diferencias=(pagos_dif / total_pagos * 100) if total_pagos > 0 else 0,
                sin_match=(pagos_sin_match / total_pagos * 100) if total_pagos > 0 else 0,
                estado='CONCILIADO' if pagos_dif == 0 and pagos_sin_match == 0 else 'CON_DIFERENCIAS',
                usuario=request.user
            )
            
            messages.success(request, f'✅ Arqueo diario generado para {hoy.strftime("%d/%m/%Y")}')
            
        except Exception as e:
            messages.error(request, f'❌ Error al generar arqueo: {str(e)}')
        
        return redirect('conciliacion')
    
    return redirect('conciliacion')
# VISTA PARA CONCILIAR PAGOS SELECCIONADOS
@login_required(login_url='/accounts/login/')
def conciliar_seleccionados(request):
    if request.method == 'POST':
        try:
            # Obtener los IDs de los pagos seleccionados
            pagos_ids = request.POST.get('pagos_ids', '')
            if pagos_ids:
                ids_lista = [int(id) for id in pagos_ids.split(',') if id.strip()]
                
                # Marcar los pagos como conciliados
                pagos_conciliados = Pago.objects.filter(id__in=ids_lista)
                for pago in pagos_conciliados:
                    pago.estado_conciliacion = 'OK'
                    pago.save()
                
                # Registrar en bitácora
                try:
                    tipo_accion, created = TipoAccion.objects.get_or_create(
                        nombre='CONCILIAR_PAGOS',
                        defaults={'estado_id': 1, 'tabla_sistema_id': 1}
                    )
                    tabla_sistema, created = TablaSistema.objects.get_or_create(
                        nombre='Conciliacion',
                        defaults={'descripcion': 'Proceso de conciliación', 'importancia': 1, 'estado_id': 1}
                    )
                    
                    Bitacora.objects.create(
                        tipo_accion=tipo_accion,
                        tabla_sistema=tabla_sistema,
                        observacion=f"Conciliados {len(ids_lista)} pagos: {', '.join([p.referencia for p in pagos_conciliados])}",
                        usuario=request.user,
                        registro_afectado=0
                    )
                except Exception as e:
                    print(f"Error al registrar en bitácora: {e}")
                
                messages.success(request, f'✅ {len(ids_lista)} pagos conciliados exitosamente.')
            else:
                messages.warning(request, 'No se seleccionaron pagos para conciliar.')
                
        except Exception as e:
            messages.error(request, f'❌ Error al conciliar pagos: {str(e)}')
        
        return redirect('conciliacion')
    
    return redirect('conciliacion')

# VISTA PARA RESOLVER DIFERENCIAS
@login_required(login_url='/accounts/login/')
def resolver_diferencia(request):
    if request.method == 'POST':
        try:
            referencia = request.POST.get('referencia')
            accion = request.POST.get('accion')
            comentario = request.POST.get('comentario', '')
            
            # Buscar el pago por referencia
            pago = Pago.objects.filter(referencia=referencia).first()
            
            if pago:
                if accion == 'ajustar_sistema':
                    # Ajustar monto en sistema interno (aquí necesitarías lógica específica)
                    pago.estado_conciliacion = 'OK'
                    pago.save()
                    mensaje = f"Ajustado monto en sistema para {referencia}"
                    
                elif accion == 'ajustar_externo':
                    # Ajustar monto en reporte externo (simulación)
                    pago.estado_conciliacion = 'OK'
                    pago.save()
                    mensaje = f"Ajustado monto en reporte externo para {referencia}"
                    
                elif accion == 'crear_incidencia':
                    # Crear incidencia automáticamente
                    ultima_incidencia = Incidencia.objects.order_by('-id').first()
                    if ultima_incidencia:
                        try:
                            ultimo_numero = int(ultima_incidencia.numero_referencia.split('-')[1])
                            nuevo_numero = f"#DIF-{ultimo_numero + 1:05d}"
                        except (ValueError, IndexError):
                            nuevo_numero = "#DIF-00001"
                    else:
                        nuevo_numero = "#DIF-00001"
                    
                    Incidencia.objects.create(
                        numero_referencia=nuevo_numero,
                        tipo_incidencia='DIF_MONTO',
                        descripcion=f"Diferencia detectada en conciliación: {referencia}. {comentario}",
                        justificacion=comentario,
                        pago_relacionado=pago,
                        usuario_asignado=request.user,
                    )
                    mensaje = f"Creada incidencia {nuevo_numero} para {referencia}"
                    
                elif accion == 'ignorar_diferencia':
                    # Ignorar diferencia y marcar como conciliado
                    pago.estado_conciliacion = 'OK'
                    pago.save()
                    mensaje = f"Diferencia ignorada para {referencia}"
                
                # Registrar en bitácora
                try:
                    tipo_accion, created = TipoAccion.objects.get_or_create(
                        nombre='RESOLVER_DIFERENCIA',
                        defaults={'estado_id': 1, 'tabla_sistema_id': 1}
                    )
                    tabla_sistema, created = TablaSistema.objects.get_or_create(
                        nombre='Conciliacion',
                        defaults={'descripcion': 'Proceso de conciliación', 'importancia': 1, 'estado_id': 1}
                    )
                    
                    Bitacora.objects.create(
                        tipo_accion=tipo_accion,
                        tabla_sistema=tabla_sistema,
                        observacion=f"{mensaje}. Acción: {accion}",
                        usuario=request.user,
                        registro_afectado=pago.id if pago else 0
                    )
                except Exception as e:
                    print(f"Error al registrar en bitácora: {e}")
                
                messages.success(request, f'✅ {mensaje}')
            else:
                messages.error(request, f'❌ No se encontró el pago con referencia {referencia}')
            
        except Exception as e:
            messages.error(request, f'❌ Error al resolver diferencia: {str(e)}')
        
        return redirect('conciliacion')
    
    return redirect('conciliacion')

# ACTUALIZA LA VISTA DE CONCILIACIÓN EXISTENTE
@login_required(login_url='/accounts/login/')
def conciliacion(request):
    # Obtener parámetros del formulario
    fecha_conciliacion = request.GET.get('fecha', '')
    pasarela_filtro = request.GET.get('pasarela', 'MacroClick')
    
    # Consulta base para pagos del sistema
    pagos_sistema = Pago.objects.all().order_by('-created_at')
    
    # Si hay fecha, filtrar por fecha
    if fecha_conciliacion:
        pagos_sistema = pagos_sistema.filter(created_at__date=fecha_conciliacion)
    
    # Simular datos externos (esto luego vendrá de archivos CSV)
    datos_externos = [
        {'referencia': '#P-00120', 'monto': 10000, 'origen': 'Tarjeta', 'estado': 'OK'},
        {'referencia': '#P-00121', 'monto': 8000, 'origen': 'Tarjeta', 'estado': 'DIF'},
        {'referencia': '', 'monto': '', 'origen': '', 'estado': 'SIN_MATCH'},
    ]
    
    # Procesar matching y estados
    for pago in pagos_sistema:
        pago.estado_conciliacion = 'SIN_MATCH'  # Por defecto
        pago.monto_externo = None
        pago.match_externo = None
        
        for externo in datos_externos:
            if externo['referencia'] == pago.referencia:
                pago.match_externo = externo
                if externo['monto'] == float(pago.monto):
                    pago.estado_conciliacion = 'OK'
                else:
                    pago.estado_conciliacion = 'DIF'
                    pago.monto_externo = externo['monto']
                    pago.diferencia_monto = abs(float(pago.monto) - externo['monto'])
                break
    
    # Calcular estadísticas mejoradas
    total_pagos = pagos_sistema.count()
    pagos_ok = sum(1 for p in pagos_sistema if getattr(p, 'estado_conciliacion', '') == 'OK')
    pagos_dif = [p for p in pagos_sistema if getattr(p, 'estado_conciliacion', '') == 'DIF']
    pagos_sin_match = [p for p in pagos_sistema if getattr(p, 'estado_conciliacion', '') == 'SIN_MATCH']
    
    # Calcular porcentajes
    porcentaje_conciliado = (pagos_ok / total_pagos * 100) if total_pagos > 0 else 0
    porcentaje_diferencias = (len(pagos_dif) / total_pagos * 100) if total_pagos > 0 else 0
    porcentaje_sin_match = (len(pagos_sin_match) / total_pagos * 100) if total_pagos > 0 else 0
    
    # Calcular total conciliado
    total_conciliado = sum(p.monto for p in pagos_sistema if getattr(p, 'estado_conciliacion', '') == 'OK')
    
    # Obtener incidencias activas recientes
    incidencias_activas = Incidencia.objects.filter(
        estado='ABIERTA'
    ).order_by('-fecha_apertura')[:5]
    
    context = {
        'pagos_sistema': pagos_sistema,
        'datos_externos': datos_externos,
        'total_pagos': total_pagos,
        'pagos_ok': pagos_ok,
        'pagos_dif': pagos_dif,
        'pagos_sin_match': pagos_sin_match,
        'porcentaje_conciliado': round(porcentaje_conciliado, 1),
        'porcentaje_diferencias': round(porcentaje_diferencias, 1),
        'porcentaje_sin_match': round(porcentaje_sin_match, 1),
        'total_conciliado': total_conciliado,
        'incidencias_activas': incidencias_activas,
        'filtros': {
            'fecha': fecha_conciliacion,
            'pasarela': pasarela_filtro,
        }
    }
    
    return render(request, 'conciliacion.html', context)